import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Workbook对象，工作簿对象，代表整个Excel文件
wb = openpyxl.load_workbook(r"hy\12\example.xlsx")
sheets = wb.sheetnames  # 获取所有工作表名称
print(sheets)  # 打印工作表名称列表

# Worksheet对象，工作表对象，代表Excel文件中的一个工作表
sheet = wb["Sheet1"]  # 获取指定工作表
print(sheet.title)  # 打印工作表名称
another_sheet = wb.active  # 获取活动工作表
print(another_sheet)  # 打印活动工作表对象

# Cell对象，单元格对象，代表工作表中的一个单元格
print(type(sheet["A1"]))  # 打印A1单元格的类型，为Cell对象
# 打印A1单元格的值，Cell对象的value属性，除此之外，还有其他属性如行row, 列column，坐标coordinate等
print(sheet["A1"].value)

# 使用表的cell方法获取单元格
print(sheet.cell(row=1, column=2))  # 获取B1单元格
# 利用cell()方法和关键字参数编写for循环
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)  # 打印B1, B3, B5, B7单元格的值

# 获取单元格最大行和列
print(sheet.max_row)  # 获取最大行数
print(sheet.max_column)  # 获取最大列数

# 列数字和字母转换
# 数字转换为字母：get_column_letter
print(get_column_letter(1))  # 获取第1列的字母
print(get_column_letter(26))  # 获取第26列的字母
# 字母转换为数字：column_index_from_string
print(column_index_from_string("A"))  # 获取A列的数字
print(column_index_from_string("Z"))  # 获取Z列的数字

# 将Worksheet对象切片，取指定区域的单元格
for row in sheet["A1":"C3"]:  # 获取A1到C3的单元格
    for cell in row:  # 遍历每一行的单元格
        print(cell.coordinate, cell.value)  # 打印单元格的值
    print()  # 换行
